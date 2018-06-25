
import os
import pymssql
import smtplib
import pymysql
import openpyxl
from datetime import date, datetime
from random import randint
from anaconda_navigator.utils.encoding import encode
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.encoders import encode_base64
from email.header import Header

'''
기본적으로 필요한 요소들.
'''

rand = randint(1, 100)
raw_now = str(datetime.now())
db_now = raw_now[0:10]
now = raw_now[0:10] + "__" + str(rand)
all_weekday = ['mon','tue','wed','thu','fri','sat','sun']
now_weekday = datetime.today().weekday()
attachments = []




'''
공통 실행함수
'''

# datetime -> string 공통변환함수
def myconverter(o):
    if isinstance(o, datetime):
        return o.__str__()


# 이메일 전송
def send_mail(from_user, to_user, subject, text, attachments):
    COMMASPACE = ", "
    msg = MIMEMultipart("alternative")
    msg["From"] = from_user
    msg["To"] = to_user
    msg["Subject"] = Header(s=subject, charset="utf-8")
    msg.attach(MIMEText(text, "html", _charset="utf-8"))

    if len(attachments) > 0:
        for attachment in attachments:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(open(attachment, "rb").read())
            encode_base64(part)
            part.add_header(
                "Content-Disposition",
                "attachment; filename=\"%s\"" % os.path.basename(attachment))
            msg.attach(part)

    server = smtplib.SMTP_SSL('smtp.worksmobile.com', 465)
    server.ehlo()
    server.login('mail address', '*******')
    server.sendmail(from_user, to_user, msg.as_string())
    server.close()




'''
그리고 실행함수들.
'''

# 바우처 계산기
def voucher_calc():

    db = pymysql.connect(
        host='hostname',
        port=3306,
        user='root',
        passwd='********',
        db='db name',
        charset='utf8')

    cursor = db.cursor()

    sql = "SELECT OES_PREFER_AREA_SIDO1, OES_PREFER_AREA_SIGUNGU1, OES_NAME, OES_HP, OES_CHILDBIRTH, OES_REGDATE FROM ORGANIZATION_ETC_SEARCHINFO WHERE OES_TYPE = 'V' AND OES_REGDATE < \'" + db_now +"\' ORDER BY OES_SEQ ASC"

    cursor.execute(sql)

    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = ['해당지사', '지역', '지역', '산모이름', '핸드폰', '출산예정일', '등록일']
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t

    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[5] = myconverter(r[5])
        if "서울특별시" in r:
            for idx, c in enumerate(r):
                if c == "강서구" or c == "양천구":
                    ws.cell(row=index + 2, column=idx).value = "강서지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "구로구" or c == "영등포구":
                    ws.cell(row=index + 2, column=idx).value = "구로지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "강북구" or c == "노원구" or c == "도봉구":
                    ws.cell(row=index + 2, column=idx).value = "노원지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "금천구" or c == "동작구" or c == "관악구":
                    ws.cell(row=index + 2, column=idx).value = "동작지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "마포구" or c == "은평구" or c == "서대문구":
                    ws.cell(row=index + 2, column=idx).value = "마포지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "성동구" or c == "광진구" or c == "중랑구":
                    ws.cell(row=index + 2, column=idx).value = "성동지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                else:
                    ws.cell(row=index + 2, column=idx + 2).value = c
        elif "경기도" in r:
            for idx, c in enumerate(r):
                if c == "구리시" or c == "남양주시":
                    ws.cell(row=index + 2, column=idx).value = "남양주지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "수원시" in c or c == "화성시":
                    ws.cell(row=index + 2, column=idx).value = "수원지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "안양시" in c or c == "의왕시" or c == "군포시" or c == "과천시":
                    ws.cell(row=index + 2, column=idx).value = "안양지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "고양시" in c or c == "김포시" or c == "파주시":
                    ws.cell(row=index + 2, column=idx).value = "일산지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                else:
                    ws.cell(row=index + 2, column=idx + 2).value = c
        else:
            for idx, c in enumerate(r):
                ws.cell(row=index + 2, column=idx + 2).value = c

    wb.save("./excel/voucher_calc_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/voucher_calc_" + now + ".xlsx")

    return True


# 바우처 계산기 - 마더앤베이비
def voucher_calc_mnb():

    db = pymysql.connect(
        host='host name',
        port=3306,
        user='root',
        passwd='********',
        db='db name',
        charset='utf8')

    cursor = db.cursor()

    sql = "SELECT OES_PREFER_AREA_SIDO1, OES_PREFER_AREA_SIGUNGU1, OES_NAME, OES_HP, OES_CHILDBIRTH, OES_REGDATE FROM ORGANIZATION_ETC_SEARCHINFO WHERE OES_TYPE = 'N' AND OES_REGDATE < \'" + db_now + "\' ORDER BY OES_SEQ ASC"

    cursor.execute(sql)


    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = ['해당지사', '지역', '지역', '산모이름', '핸드폰', '출산예정일', '등록일']
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t

    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[5] = myconverter(r[5])
        if "서울특별시" in r:
            for idx, c in enumerate(r):
                if c == "강서구" or c == "양천구":
                    ws.cell(row=index + 2, column=idx).value = "강서지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "구로구" or c == "영등포구":
                    ws.cell(row=index + 2, column=idx).value = "구로지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "강북구" or c == "노원구" or c == "도봉구":
                    ws.cell(row=index + 2, column=idx).value = "노원지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "금천구" or c == "동작구" or c == "관악구":
                    ws.cell(row=index + 2, column=idx).value = "동작지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "마포구" or c == "은평구" or c == "서대문구":
                    ws.cell(row=index + 2, column=idx).value = "마포지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "성동구" or c == "광진구" or c == "중랑구":
                    ws.cell(row=index + 2, column=idx).value = "성동지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                else:
                    ws.cell(row=index + 2, column=idx + 2).value = c
        elif "경기도" in r:
            for idx, c in enumerate(r):
                if c == "구리시" or c == "남양주시":
                    ws.cell(row=index + 2, column=idx).value = "남양주지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "수원시" in c or c == "화성시":
                    ws.cell(row=index + 2, column=idx).value = "수원지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "안양시" in c or c == "의왕시" or c == "군포시" or c == "과천시":
                    ws.cell(row=index + 2, column=idx).value = "안양지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "고양시" in c or c == "김포시" or c == "파주시":
                    ws.cell(row=index + 2, column=idx).value = "일산지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                else:
                    ws.cell(row=index + 2, column=idx + 2).value = c
        else:
            for idx, c in enumerate(r):
                ws.cell(row=index + 2, column=idx + 2).value = c

    # 엑셀 파일 저장
    wb.save("./excel/voucher_calc_mnb_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/voucher_calc_mnb_" + now + ".xlsx")
    return True


# 맘초회원
def momcho_member():

    db = pymysql.connect(
        host='host name',
        port=3306,
        user='root',
        passwd='********',
        db='db name',
        charset='utf8')

    cursor = db.cursor()

    sql = "SELECT MU_NAME, MU_MOBILE, MU_CHILD_BIRTH, MU_REGDATE, MU_BIRTH, MU_EMAIL,MU_GENDER,  MU_CENTER_INFO_FLAG FROM MEMBER_USER WHERE MU_CENTER_INFO_FLAG = 1 ORDER BY MU_REGDATE ASC"

    cursor.execute(sql)

    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = ['산모이름', '핸드폰번호', '출산예정일', '등록일', '산모생일', '산모이메일', '성별','조리원 정보제공 동의여부']
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t

    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[3] = myconverter(r[3])
        r[7] = int.from_bytes(r[7], byteorder='big')
        # print(r)

        for idx, c in enumerate(r):
            # print(c)
            if c == 1:
                ws.cell(row=index+2, column=idx+1).value = "동의"
            else:
                ws.cell(row=index+2, column=idx+1).value = c



    # 엑셀 파일 저장
    wb.save("./excel/momcho_member_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/momcho_member_" + now + ".xlsx")

    return True


# 맘초 마더앤베이비 상담
def momcho_mnb():
    ## 해당 지사 분리
    db = pymysql.connect(
        host='host name',
        port=3306,
        user='root',
        passwd='*******',
        db='db name',
        charset='utf8')

    cursor = db.cursor()

    sql = "SELECT OES_PREFER_AREA_SIDO1, OES_PREFER_AREA_SIGUNGU1,OES_NAME, OES_HP ,OES_CHILDBIRTH, OES_REGDATE  FROM ORGANIZATION_ETC_SEARCHINFO WHERE OES_TYPE = 'M' AND OES_CENTER_INFO_FLAG = 1 ORDER BY OES_SEQ ASC"

    cursor.execute(sql)

    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = ['해당지사','지역', '지역', '산모이름', '핸드폰번호','출산예정일', '등록일']
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t

    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[5] = myconverter(r[5])

        # r[6] = int.from_bytes(r[6], byteorder='big')
        # for idx, c in enumerate(r):
        # if c == 1:
        #     ws.cell(row=index+2, column=idx+1).value = "동의"
        # else:
        #     ws.cell(row=index+2, column=idx+1).value = c

        # print(r)

        if "서울특별시" in r:
            for idx, c in enumerate(r):
                if c == "강서구" or c == "양천구":
                    ws.cell(row=index + 2, column=idx).value = "강서지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "구로구" or c == "영등포구":
                    ws.cell(row=index + 2, column=idx).value = "구로지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "강북구" or c == "노원구" or c == "도봉구":
                    ws.cell(row=index + 2, column=idx).value = "노원지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "금천구" or c == "동작구" or c == "관악구":
                    ws.cell(row=index + 2, column=idx).value = "동작지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "마포구" or c == "은평구" or c == "서대문구":
                    ws.cell(row=index + 2, column=idx).value = "마포지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif c == "성동구" or c == "광진구" or c == "중랑구":
                    ws.cell(row=index + 2, column=idx).value = "성동지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                else:
                    ws.cell(row=index + 2, column=idx + 2).value = c
        elif "경기도" in r:
            for idx, c in enumerate(r):
                if c == "구리시" or c == "남양주시":
                    ws.cell(row=index + 2, column=idx).value = "남양주지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "수원시" in c or c == "화성시":
                    ws.cell(row=index + 2, column=idx).value = "수원지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "안양시" in c or c == "의왕시" or c == "군포시" or c == "과천시":
                    ws.cell(row=index + 2, column=idx).value = "안양지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                elif "고양시" in c or c == "김포시" or c == "파주시":
                    ws.cell(row=index + 2, column=idx).value = "일산지사"
                    ws.cell(row=index + 2, column=idx + 2).value = c
                else:
                    ws.cell(row=index + 2, column=idx + 2).value = c
        else:
            for idx, c in enumerate(r):
                ws.cell(row=index + 2, column=idx + 2).value = c
    # 엑셀 파일 저장
    wb.save("./excel/momcho_mnb_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/momcho_mnb_" + now + ".xlsx")
    return True


# 간편견적 디비
def simple_consult():

    db = pymysql.connect(
        host='host name',
        port=3306,
        user='root',
        passwd='******',
        db='db name',
        charset='utf8')

    cursor = db.cursor()

    sql = "SELECT OES_PREFER_AREA_SIDO1, OES_PREFER_AREA_SIGUNGU1, OES_NAME, OES_HP, OES_CHILDBIRTH, OES_REGDATE, COMMON_CODE_DETAIL_NAME(OES_SCHEDULE_CHILD), OES_CENTER_INFO_FLAG FROM ORGANIZATION_ETC_SEARCHINFO WHERE OES_TYPE = 'S' ORDER BY OES_SEQ ASC "

    cursor.execute(sql)

    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = ['지역', '지역', '산모이름', '핸드폰번호', '출산예정일', '등록일', '출산자녀','제3자 정보제공동의여부']
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t

    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[5] = myconverter(r[5])
        r[7] = int.from_bytes(r[7], byteorder='big')
        for idx, c in enumerate(r):
            if c == 1:
                ws.cell(row=index + 2, column=idx + 1).value = "동의"
            else:
                ws.cell(row=index + 2, column=idx + 1).value = c

    # 엑셀 파일 저장
    wb.save("./excel/simple_consult_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/simple_consult_" + now + ".xlsx")

    return True


# 마더앤베이비 가입회원 - 산후조리원 정보제공 약관 수정이후 가입
def mnb_member():

    db = pymssql.connect(
        server='host name',
        user='user name',
        password='********',
        database='db name'
        )

    cursor = db.cursor()


    sql = "SELECT ADDINFO.MD_ADDR1, BA.MM_MEM_NM, ADDINFO.MD_HP, ADDINFO.MD_CHILDBIRTH, BA.MM_ENT_DT FROM [dbo].[MO_MEM_MAS] AS BA INNER JOIN [dbo].[MO_MEM_DTL] AS ADDINFO ON BA.MM_MEM_ID = ADDINFO.MD_MEM_ID WHERE BA.MM_ENT_DT > '2018-05-17' ORDER BY BA.MM_ENT_DT ASC"

    cursor.execute(sql)

    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = [
        '지역', '산모이름', '핸드폰번호', '출산예정일', '등록일'
    ]
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t


    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[4] = myconverter(r[4])
        for idx, c in enumerate(r):
            addr = r[0].encode('ISO-8859-1')
            addr = addr.decode('ms949')
            name = r[1].encode('ISO-8859-1')
            name = name.decode('ms949')
            if idx==0:
                ws.cell(row=index + 2, column=idx + 1).value = addr
            elif idx==1:
                ws.cell(row=index + 2, column=idx + 1).value = name
            else:
                ws.cell(row=index + 2, column=idx + 1).value = c


    # 엑셀 파일 저장
    wb.save("./excel/mnb_member_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/mnb_member_" + now + ".xlsx")

    return True

# 폴스베이비에 전달할 디비
def pauls_baby():

    db = pymssql.connect(
        server='host name',
        user='user name',
        password='********',
        database='db name')

    cursor = db.cursor()

    sql = "SELECT MR_ADDR1, MR_ADDR2, MR_MEM_NM, MR_HP, CONVERT(VARCHAR(10), CONVERT(DATETIME, MR_CHILDBIRTH),121) AS MR_CHILDBIRTH, MR_REG_DT, MR_PAY_TYPE, MR_PAY_STATE FROM [dbo].[MO_RESERVATION] WHERE MR_CHILDBIRTH >  \'" + db_now + "\'  AND MR_CANCEL_YN = 'N' ORDER BY MR_CHILDBIRTH"

    cursor.execute(sql)

    data = cursor.fetchall()

    data = list(data)

    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    # 엑셀파일에 타이틀 넣기
    title = ['지역', '지역','산모이름', '핸드폰번호', '출산예정일', '등록일', '결제방법','결제상태']
    for index, t in enumerate(title):
        ws.cell(row=1, column=index + 1).value = t

    # DB 추출한 데이터 행마다 넣기
    for index, row in enumerate(data):
        r = list(data[index])
        r[5] = myconverter(r[5])
        for idx, c in enumerate(r):
            addr1 = r[0].encode('ISO-8859-1')
            addr1 = addr1.decode('ms949')
            addr2 = r[1].encode('ISO-8859-1')
            addr2 = addr2.decode('ms949')
            name = r[2].encode('ISO-8859-1')
            name = name.decode('ms949')
            pay_type = r[6].encode('ISO-8859-1')
            pay_type = pay_type.decode('ms949')
            pay_state = r[7].encode('ISO-8859-1')
            pay_state = pay_state.decode('ms949')
            if idx == 0:
                ws.cell(row=index + 2, column=idx + 1).value = addr1
            elif idx == 1:
                ws.cell(row=index + 2, column=idx + 1).value = addr2
            elif idx == 2:
                ws.cell(row=index + 2, column=idx + 1).value = name
            elif idx == 6:
                ws.cell(row=index + 2, column=idx + 1).value = pay_type
            elif idx == 7:
                ws.cell(row=index + 2, column=idx + 1).value = pay_state
            else:
                ws.cell(row=index + 2, column=idx + 1).value = c

    # 엑셀 파일 저장
    wb.save("./excel/pauls_baby_" + now + ".xlsx")
    wb.close()

    attachments.append("./excel/pauls_baby_" + now + ".xlsx")

    return True




to = 'mail address'
sender = 'mail address'

# 정리된 디비 메일 뿌리기
voucher_calc()
voucher_calc_mnb()
momcho_member()
momcho_mnb()
simple_consult()
mnb_member()
if all_weekday[now_weekday] == 'fri': 
    pauls_baby()

send_mail(sender, to, "디비정리파일",
          "안녕하세요. ***입니다. 업무에 필요한 파일 정리하여 첨부합니다.", attachments)



import sys
sys.exit()  # 프로세스 끝내기
